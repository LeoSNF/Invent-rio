import pandas as pd
import math
# from datetime import datetime, timedelta


def processar_mb52(path):
    """
    Processa o arquivo MB52.xlsx e retorna DataFrame formatado
    """
    cols = [
        "Depósito", "Centro", "Material", "Texto breve material", "Lote",
        "Utilização livre", "Controle qualidade", "Bloqueado", "Estq.não disponível",
        "Trânsito e TE", "UM básica", "Val.utiliz.livre", "Valor verif.qual.",
        "Val.estoque bloq.", "Val.util.restrita", "Val.em trâns.e Trf", "Valor em trânsito"
    ]
    
    try:
        df = pd.read_excel(path, usecols=cols)
        df["Material"] = df["Material"].apply(lambda x: str(int(x)) if pd.notnull(x) and str(x).endswith('.0') else str(x))
        df["Concatenado"] = df["Material"].str.replace('.0', '', regex=False) + df["Lote"].astype(str)
        return df
    except Exception as e:
        print(f"Erro ao processar MB52: {e}")
        return pd.DataFrame()

def merge_bmbc(df, bmbc_path):
    """
    Faz merge do DataFrame principal com dados do BMBC
    """
    try:
        df_bmbc = pd.read_excel(bmbc_path)
        col_material = next((col for col in df_bmbc.columns if "material" in col.lower()), None)
        col_lote = next((col for col in df_bmbc.columns if "lote" in col.lower()), None)
        col_validade = next((col for col in df_bmbc.columns if "venc" in col.lower() or "valid" in col.lower()), None)
        
        if not (col_material and col_lote and col_validade):
            return df
        
        df_bmbc[col_material] = df_bmbc[col_material].apply(lambda x: str(int(x)) if pd.notnull(x) and str(x).endswith('.0') else str(x))
        df_bmbc[col_lote] = df_bmbc[col_lote].astype(str)
        
        df_merged = pd.merge(
            df,
            df_bmbc[[col_material, col_lote, col_validade]],
            left_on=["Material", "Lote"],
            right_on=[col_material, col_lote],
            how="left"
        )
        
        df_merged.rename(columns={col_validade: "Validade SAP"}, inplace=True)
        
        if col_material != "Material":
            df_merged.drop(columns=[col_material], inplace=True)
        if col_lote != "Lote":
            df_merged.drop(columns=[col_lote], inplace=True)
            
        return df_merged
    except Exception as e:
        print(f"Erro ao fazer merge com BMBC: {e}")
        return df

def merge_inventory(df, inventory_path):
    """
    Faz merge do DataFrame principal com dados do Inventory
    """
    try:
        df_inventory = pd.read_excel(inventory_path)
        df['Material'] = df['Material'].astype(str).str.strip().str.upper()
        df['Texto breve material'] = df['Texto breve material'].astype(str).str.strip().str.upper()
        df_inventory['NUMERO'] = df_inventory['NUMERO'].astype(str).str.strip().str.upper()
        df_inventory['DESCRICAO'] = df_inventory['DESCRICAO'].astype(str).str.strip().str.upper()
        
        df_merged = pd.merge(
            df,
            df_inventory[['NUMERO', 'DESCRICAO', 'DATA']],
            left_on=['Material', 'Texto breve material'],
            right_on=['NUMERO', 'DESCRICAO'],
            how="left"
        )
        
        df_merged.drop(columns=['NUMERO', 'DESCRICAO'], inplace=True)
        return df_merged
    except Exception as e:
        print(f"Erro ao fazer merge com Inventory: {e}")
        return df

def merge_reconciliation(df, reconciliation_path):
    """
    Faz merge do DataFrame principal com dados de reconciliação
    """
    try:
        df_reconciliation = pd.read_excel(reconciliation_path)
        df['Material'] = df['Material'].astype(str).str.strip().str.upper()
        df_reconciliation['Material'] = df_reconciliation['Material'].astype(str).str.strip().str.upper()
        
        df_merged = pd.merge(
            df,
            df_reconciliation[['Material', 'Validade NTV']],
            on='Material',
            how="left"
        )
        
        return df_merged
    except Exception as e:
        print(f"Erro ao fazer merge com Reconciliation: {e}")
        return df

def formatar_datas(df, colunas_data):
    """
    Formata colunas de data no DataFrame
    """
    try:
        for coluna in colunas_data:
            if coluna in df.columns:
                df[coluna] = pd.to_datetime(df[coluna], errors='coerce')
                df[coluna] = df[coluna].dt.strftime('%d/%m/%Y')
        return df
    except Exception as e:
        print(f"Erro ao formatar datas: {e}")
        return df

def remover_duplicatas(df):
    """
    Remove linhas duplicadas do DataFrame
    """
    try:
        return df.drop_duplicates()
    except Exception as e:
        print(f"Erro ao remover duplicatas: {e}")
        return df

def dias_para_meses(dias):
    """
    Converte dias para meses
    """
    try:
        if pd.isna(dias) or dias == 0:
            return 0
        return math.ceil(dias / 30.44)  # Média de dias por mês
    except:
        return 0

def validar_vencimento(data_validade, dias_limite=30):
    """
    Valida se um item está próximo do vencimento
    """
    try:
        if pd.isna(data_validade):
            return False
        
        data_validade = pd.to_datetime(data_validade)
        hoje = pd.Timestamp.now()
        dias_restantes = (data_validade - hoje).days
        
        return dias_restantes <= dias_limite
    except:
        return False

def calcular_age_limit(material, one_portfolio_data):
    """
    Calcula o age limit para um material baseado no One Portfolio
    """
    try:
        if material in one_portfolio_data:
            return one_portfolio_data[material]
        return None
    except:
        return None

def aplicar_formatacao_excel(df, save_path):
    """
    Aplica formatação ao arquivo Excel
    """
    try:
        # Salvar DataFrame
        df.to_excel(save_path, index=False)
        
        # Carregar workbook para formatação
        import openpyxl
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import PatternFill, Font, Alignment
        
        wb = openpyxl.load_workbook(save_path)
        ws = wb.worksheets[0]
        
        # Configurar tabela
        max_row = ws.max_row
        max_col = ws.max_column
        last_col_letter = get_column_letter(max_col)
        table_range = f"A1:{last_col_letter}{max_row}"
        tab = Table(displayName="ControleInventario", ref=table_range)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        
        # Formatar cabeçalho
        for cell in ws["1"]:
            cell.fill = PatternFill(start_color="C2B7BD", end_color="C2B7BD", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Aplicar validações de vencimento
        if "Validade SAP" in df.columns:
            sap_col = df.columns.get_loc("Validade SAP") + 1
            for row in range(2, max_row + 1):
                cell = ws.cell(row=row, column=sap_col)
                if cell.value:
                    try:
                        data_validade = pd.to_datetime(cell.value)
                        hoje = pd.Timestamp.now()
                        dias_restantes = (data_validade - hoje).days
                        
                        if dias_restantes <= 30:
                            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                            cell.font = Font(color="FFFFFF", bold=True)
                        elif dias_restantes <= 90:
                            cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                            cell.font = Font(color="FFFFFF", bold=True)
                    except:
                        pass
        
        wb.save(save_path)
        return True
    except Exception as e:
        print(f"Erro ao aplicar formatação Excel: {e}")
        return False

def verificar_arquivo_existe(caminho):
    """
    Verifica se um arquivo existe
    """
    import os
    return os.path.exists(caminho)

def obter_info_arquivo(caminho):
    """
    Obtém informações básicas de um arquivo
    """
    try:
        if not verificar_arquivo_existe(caminho):
            return None
        
        df = pd.read_excel(caminho)
        return {
            'linhas': len(df),
            'colunas': len(df.columns),
            'nome_colunas': list(df.columns)
        }
    except Exception as e:
        print(f"Erro ao obter informações do arquivo: {e}")
        return None 