from modulos.sap import Sap
from modulos.formInventory import FormInventory, CommandsCodes
from modulos.utils import processar_mb52, merge_bmbc, merge_inventory, merge_reconciliation, formatar_datas, remover_duplicatas, dias_para_meses
from modulos.msgBox import MsgBox, MsgBoxOptions
from modulos.parameters import Parameters

import ttkbootstrap as tb
import os
import time
import pandas as pd
from io import StringIO
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
import logging

logging.basicConfig(format='%(asctime)s - %(levelname)s: %(message)s',
                    filename=f'{os.getcwd()}\\log.txt', filemode='w',
                    datefmt='%d-%b-%y %H:%M:%S', level=logging.INFO,
                    encoding='utf-8')


class Inventory(Sap):
    def __init__(self) -> None:
        p = Parameters(b'pVOxw0erfHGl4agvXg-nQlu6PySZYn6m7-_kZrlJ3yQ=')
        p.readParameters()

        self.__scriptName = os.path.basename(__file__).replace('.py','')
        self.execFail = False
        self.rootPath = os.getcwd()
        self.save_path = self.rootPath + '\\Controle de Inventário.xlsx'
        self.__msgBox = MsgBox()
        
        # self.mhtmlSapPth = self.rootPath
        # sapEnv = 'PHS [sapphsas01.pharma.aventis.com]' # Parametrizar
        # sapLang = 'EN' # Parametrizar
        # connectBy = 2 # Parametrizar
        # sapFile = 'C:\\Program Files\\SAP\\FrontEnd\\SapGui\\saplogon.exe' # Parametrizar

        sapEnv = p.sapEnv
        sapLang = p.sapLang
        connectBy = p.sapConnBy
        sapFile = p.sapFile
        newSession = bool(p.sapNewSession)
        self.mhtmlSapPth = p.sapExportFileDestPath if p.sapExportFileDestPath else self.rootPath

        try:
            super().__init__(sapEnv, userId='', userPW='', language=sapLang, connectBy=connectBy, sapFile=sapFile, newSession=newSession)
            # pass
        except Exception as e:
            print(str(e))
            logging.error(f'__init__:\n\n {str(e)}')
            self.__msgBox.showMsgBox(MsgBoxOptions.ERROR, self.__scriptName, 'Falha ao iniciar o SAP.')
            self.execFail = True

    def __connectSap(self) -> None:
        super().openSap()

    def __showFormAndReturnActionCode(self) -> tuple:
        try:
            root = tb.Window(themename="flatly")
            root.wm_attributes("-topmost", 1)
            frmInvt = FormInventory(root)
            root.mainloop()
            actionOpt:int = frmInvt.commandCode
            mb52FilePth:str = frmInvt.mb52_path
            bmbcFilePth:str = frmInvt.bmbc_path
            inventoryFilePth:str = frmInvt.inventory_path
            onePortifolioFilePth:str = frmInvt.one_portfolio_path
            
            del frmInvt
            del root

            return (actionOpt, mb52FilePth, bmbcFilePth, inventoryFilePth, onePortifolioFilePth)
        except Exception as e:
            logging.error(f'__showFormAndReturnActionCode:\n\n {str(e)}')
            self.execFail = True
            return (None, None, None, None, None, None)
        

    def run(self) -> None:            
        while True:
            successMsg = ''
            askToCloseExcel = False
            actionCode, mb52FilePth, bmbcFilePth, inventoryFilePth, onePortifolioFilePth = self.__showFormAndReturnActionCode()

            if self.execFail:
                self.__msgBox.showMsgBox(MsgBoxOptions.ERROR, self.__scriptName, 'Falha no processamento do formulário.')
                return

            print(f'Action code: {actionCode}\nmb52FilePthfile path: {mb52FilePth}\nbmbcFilePth file path: {bmbcFilePth}\nonePortifolioFilePth file path: {onePortifolioFilePth}\ninventoryFilePth file path: {inventoryFilePth}')

            match actionCode:
                case CommandsCodes.MB52.value:
                    successMsg = self.__importDataFromMb52()
                    successMsg += self.__converter_mhtml_para_excel(self.mhtmlSapPth + '\\MB52.MHTML', self.mhtmlSapPth + '\\MB52.xlsx')
                    if not self.execFail: askToCloseExcel = True
                case CommandsCodes.BMBC.value:
                    successMsg = self.__importDataFromBmbc(False)
                    successMsg += self.__converter_mhtml_para_excel(self.mhtmlSapPth + '\\BMBC.MHTML', self.mhtmlSapPth + '\\BMBC.xlsx')
                    if not self.execFail: askToCloseExcel = True
                case CommandsCodes.CALCINVT.value:
                    self.__validateInputFilesPaths(mb52FilePth, bmbcFilePth, inventoryFilePth, onePortifolioFilePth)
                    successMsg = self.__processInventory(mb52FilePth, bmbcFilePth, inventoryFilePth, onePortifolioFilePth)
                case CommandsCodes.CANCEL.value:
                    print('Processo cancelado pelo usuário.')
                    logging.info('Run: Processo cancelado pelo usuário.')
                    return
                case _:
                    print('Unknown Inventory form option')
                    logging.error('Unknown Inventory form option')
                    self.__msgBox.showMsgBox(MsgBoxOptions.ERROR, self.__scriptName, 'Falha no processamento do formulário.')
                    return
            
            if askToCloseExcel:
                self.__msgBox.showMsgBox(MsgBoxOptions.WARNINGOPT, self.__scriptName, 'Deseja fechar TODAS as planilhas do Excel abertas?' \
                                                                                      '\nATENÇÃO! Planilhas aberta que não foram salvas terão seu conteúdo perdido.')
                if self.__msgBox.answer: self.__closeExcelForced()
            
            if self.execFail:
                self.__msgBox.showMsgBox(MsgBoxOptions.ERROR, self.__scriptName, 'Falha na execução da ferramenta.')
            else:
                self.__msgBox.showMsgBox(MsgBoxOptions.INFO, self.__scriptName, successMsg)
                logging.info('Run: Processo finalizado com sucesso.')
            
            self.execFail = False
            
    def __importDataFromMb52(self) -> str:
        """
        Executa a transação MB52
        """

        try:
            print('Importando arquivos de MB52')

            self.__connectSap()
            self.executeTransaction("MB52")

            self.session.findById("wnd[0]/tbar[1]/btn[17]").press()
            self.session.findById("wnd[1]/usr/txtV-LOW").text = "CONTROLE_INV"
            self.session.findById("wnd[1]/usr/txtENAME-LOW").text = ""
            self.session.findById("wnd[0]/tbar[1]/btn[8]").press()
            self.session.findById("wnd[0]/tbar[1]/btn[8]").press()
            
            # Move o arquivo para a pasta de destino
            self.session.findById("wnd[0]/tbar[1]/btn[43]").press() 
            self.session.findById("wnd[1]/usr/ctxtDY_PATH").text = self.mhtmlSapPth # r"C:\Users\i0336821\OneDrive - Sanofi\Área de Trabalho\Projetos\Inventário"
            self.session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "MB52.MHTML"
            self.session.findById("wnd[1]/tbar[0]/btn[11]").press()
            # self.session.findById("wnd[0]").close()
            # self.session.findById("wnd[1]/usr/btnSPOP-OPTION1").press()

            return f'Execução da transação MB52 realizada com sucesso! Arquivo enviado para: {self.mhtmlSapPth}'

        except Exception as e:
            self.execFail = True
            print(str(e))
            logging.error(f'__importDataFromMb52:\n\n {str(e)}')
            return ''
            
    def __importDataFromBmbc(self, colar_material=True) -> str:
        """
        Executa a transação BMBC com opção de colar dados da coluna Material
        
        Args:
            session: Sessão SAP
            colar_material (bool): Se True, cola os dados da área de transferência
        """
        try:
            # Navegação e exportação conforme estrutura fornecida
            print('Importando arquivos de BMBC')

            self.__connectSap()
            self.executeTransaction("BMBC")

            materials:pd.DataFrame = self.__getMaterialCodesFromMb52()

            materials.to_clipboard(index=False, header=False)

            self.multipleSelection("wnd[0]/usr/subLIMITATIONS:RVBBINCO:0101/subSUB:RVBBINCO:0105/tabsTABSTRIP/tabpPUSH01/ssubLIMITATIONS:RVBBINCO:0106/subSELECTION:RVBBINCO:0140/subSUB:SAPLSSEL:1106/btn%_%%DYN001_%_APP_%-VALU_PUSH")

            # self.session.findById("wnd[0]/usr/subLIMITATIONS:RVBBINCO:0101/subSUB:RVBBINCO:0105/tabsTABSTRIP/tabpPUSH01/ssubLIMITATIONS:RVBBINCO:0106/subSELECTION:RVBBINCO:0110/subSUB:SAPLSSEL:1106/btn%_%%DYN001_%_APP_%-VALU_PUSH").press()
            
            # Se deve colar os dados da área de transferência

            # if colar_material:
            #     print("Colando dados da coluna Material na tela de seleção...")
            #     # Aguarda um pouco para a tela carregar
            #     time.sleep(2)
            #     # Cola os dados da área de transferência (Ctrl+V)
            #     self.session.findById("wnd[1]").sendVKey(2)  # VKey 2 = Ctrl+V
            #     time.sleep(1)
            self.session.findById("wnd[0]/tbar[1]/btn[8]").press()
            self.session.findById("wnd[0]/shellcont/shell/shellcont[0]/shell/shellcont[0]/shell/shellcont[1]/shell[0]").pressButton("MAST_GRID")
            self.session.findById("wnd[0]/shellcont/shell/shellcont[0]/shell/shellcont[0]/shell").pressToolbarContextButton("&MB_EXPORT")
            self.session.findById("wnd[0]/shellcont/shell/shellcont[0]/shell/shellcont[0]/shell").selectContextMenuItem("&XXL")
            self.session.findById("wnd[1]/usr/ctxtDY_PATH").text = self.mhtmlSapPth
            self.session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "BMBC.MHTML"
            self.session.findById("wnd[1]/tbar[0]/btn[11]").press()

            return f'Execução da transação BMBC realizada com sucesso!Arquivo enviado para: {self.mhtmlSapPth}'
        
        except Exception as e:
            self.execFail = True
            print(str(e))
            logging.error(f'__importDataFromBmbc:\n\n {str(e)}')
            return ''

    def __getMaterialCodesFromMb52(self) -> pd.DataFrame:
        mb52XlsxPath = self.mhtmlSapPth + '\\MB52.xlsx'

        if not os.path.exists(mb52XlsxPath):
            print(f"Arquivo {mb52XlsxPath} não encontrado")
            raise FileNotFoundError(f"Arquivo {mb52XlsxPath} não encontrado")
        
        df = pd.read_excel(mb52XlsxPath)

        df.drop_duplicates(subset=['Material'], inplace=True)

        return df[~df['Material'].isna()]['Material'].astype(str).str.replace('.0', '').copy()

    def __converter_mhtml_para_excel(self, arquivo_mhtml:str, arquivo_excel:str):
        try:
            if not os.path.exists(arquivo_mhtml): 
                print(f"Arquivo {arquivo_mhtml} não encontrado")
                raise FileNotFoundError(f"Arquivo {arquivo_mhtml} não encontrado")
                
            if arquivo_excel is None:
                arquivo_excel = os.path.splitext(arquivo_mhtml)[0] + ".xlsx"
                
            with open(arquivo_mhtml, "r", encoding="utf-8") as file:
                conteudo = file.read()
                
            tabelas = pd.read_html(StringIO(conteudo), header=0)
                
            if len(tabelas) > 0:
                df = tabelas[0]
                df.to_excel(arquivo_excel, index=False)
                print(f"Arquivo Excel gerado: {arquivo_excel}")
                return f"Arquivo Excel gerado: {arquivo_excel}"
            else:
                print("Nenhuma tabela encontrada no arquivo MHTML")
                raise ValueError("Nenhuma tabela encontrada no arquivo MHTML")
    
        except Exception as e:
            self.execFail = True
            print(f"Erro ao converter MHTML para Excel:\n\n{str(e)}")
            logging.error(f"Erro ao converter MHTML para Excel:\n\n{str(e)}")
            return ''

    def __closeExcelForced(self):
        time.sleep(1)
        os.system('taskkill /f /im excel.exe')
        time.sleep(2)

    def __validateInputFilesPaths(self, mb52Path:str, bmbcPath:str, inventoryPath:str, onePortfolioPath:str) -> str:
        errMsg = ''


        if mb52Path.split('\\')[-1].split('.')[-1].lower() != 'xlsx':
            errMsg += '- Extensão do arquivo MB52 incorreto\n'
        
        if bmbcPath.split('\\')[-1].split('.')[-1].lower() != 'xlsx':
            errMsg += '- Extensão do arquivo BMBC incorreto\n'

        if inventoryPath.split('\\')[-1].split('.')[-1].lower() != 'xlsx':
            errMsg += '- Extensão do arquivo de inventório incorreto\n'

        if onePortfolioPath.split('\\')[-1].split('.')[-1].lower() != 'xlsb':
            errMsg += '- Extensão do arquivo Oneportifolio incorreto\n'
        
        if errMsg:
            errMsg = 'Atenção!\n\n' + errMsg
        
        return errMsg


    def __processInventory(self, mb52Path:str, bmbcPath:str, inventoryPath:str, onePortfolioPath:str) -> str:
        # save_path = "Controle de Inventário.xlsx"
        print('Processando inventório...')
        
        try:
            df = processar_mb52(mb52Path)
            if bmbcPath:
                df = merge_bmbc(df, bmbcPath)
            if inventoryPath:
                df = merge_inventory(df, inventoryPath)
            ordem_colunas = [
                "Depósito", "Centro", "Material", "Texto breve material", "Lote",
                "Utilização livre", "Controle qualidade", "Bloqueado", "Estq.não disponível",
                "Trânsito e TE", "UM básica", "Val.utiliz.livre", "Valor verif.qual.",
                "Val.estoque bloq.", "Val.util.restrita", "Val.em trâns.e Trf", "Valor em trânsito",
                "Concatenado"
            ]
            if "Validade SAP" in df.columns:
                ordem_colunas.append("Validade SAP")
            if "Validade NTV" in df.columns:
                ordem_colunas.append("Validade NTV")
            if "DATA" in df.columns:
                ordem_colunas.append("DATA")
            df = df[ordem_colunas] if all(col in df.columns for col in ordem_colunas) else df
            df = remover_duplicatas(df)
            if "Validade SAP" in df.columns:
                df = formatar_datas(df, ["Validade SAP"])
            if "Validade NTV" in df.columns:
                df = formatar_datas(df, ["Validade NTV"])
            if "DATA" in df.columns:
                df = formatar_datas(df, ["DATA"])
            
            # Salvar arquivo
            self.__salvar_e_formatar_excel(df, self.save_path)
            
            # Merge com One Portfolio se disponível
            if onePortfolioPath:
                try:
                    # df_controle = pd.read_excel(save_path)
                    df_controle = df.copy()
                    df_portfolio = pd.read_excel(onePortfolioPath, sheet_name='VW_S&OP', header=5)
                    df_portfolio['GMID'] = df_portfolio['GMID'].astype(str).str.strip().str.upper()
                    df_controle['Material'] = df_controle['Material'].astype(str).str.strip().str.upper()
                    df_merged_port = pd.merge(
                        df_controle,
                        df_portfolio[['GMID', 'Age Limit (months)']],
                        left_on='Material',
                        right_on='GMID',
                        how="left"
                    )
                    df_merged_port.rename(columns={'Age Limit (months)': 'Age Limit_SAP'}, inplace=True)
                    if 'GMID' in df_merged_port.columns:
                        df_merged_port.drop(columns=['GMID'], inplace=True)
                    df_merged_port.to_excel(self.save_path, index=False)
                    # Reaplica formatação após merge com One Portfolio
                    wb = openpyxl.load_workbook(self.save_path)
                    ws = wb.worksheets[0]
                    max_row = ws.max_row
                    max_col = ws.max_column
                    last_col_letter = get_column_letter(max_col)
                    table_range = f"A1:{last_col_letter}{max_row}"
                    tab = Table(displayName="ControleInventario", ref=table_range)
                    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
                    tab.tableStyleInfo = style
                    ws.add_table(tab)
                    for cell in ws["1"]:
                        cell.fill = PatternFill(start_color="C2B7BD", end_color="C2B7BD", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                    wb.save(self.save_path)
                except Exception as e:
                    print(f"Erro ao processar One Portfolio: {e}")
                    logging.error(f'__processInventory: Erro ao processar One Portfolio \n\n{str(e)}')
                
            # messagebox.showinfo("Sucesso", f"Inventário processado com sucesso!\nArquivo salvo como: {save_path}")
            logging.info(f"Inventário processado com sucesso!\nArquivo salvo como: {self.save_path}")
            return f"Inventário processado com sucesso!\nArquivo salvo como: {self.save_path}"

        except Exception as e:
            self.execFail = True
            logging.error(f'__processInventory:\n\n{str(e)}')
            # messagebox.showerror("Erro", f"Erro ao processar inventário: {str(e)}")
            return ''

    def __salvar_e_formatar_excel(self, df:pd.DataFrame, save_path:str):
        # Salvar DataFrame
        df.to_excel(save_path, index=False)
        
        # Carregar workbook para formatação
        wb = openpyxl.load_workbook(save_path)
        ws = wb.worksheets[0]
        
        # Configurar tabela
        max_row = ws.max_row
        max_col = ws.max_column
        last_col_letter = get_column_letter(max_col)
        table_range = f"A1:{last_col_letter}{max_row}"
        tab = Table(displayName="ControleInventario", ref=table_range)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        
        # Formatar cabeçalho
        for cell in ws["1"]:
            cell.fill = PatternFill(start_color="C2B7BD", end_color="C2B7BD", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Aplicar validações
        if "Validade SAP" in df.columns:
            sap_col = df.columns.get_loc("Validade SAP") + 1
            for row in range(2, max_row + 1):
                cell = ws.cell(row=row, column=sap_col)
                if cell.value:
                    try:
                        data_validade = pd.to_datetime(cell.value)
                        hoje = pd.Timestamp.now()
                        dias_restantes = (data_validade - hoje).days
                        
                        if dias_restantes <= 30:
                            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                            cell.font = Font(color="FFFFFF", bold=True)
                        elif dias_restantes <= 90:
                            cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                            cell.font = Font(color="FFFFFF", bold=True)
                    except:
                        pass
        
        wb.save(save_path)    

invObj = Inventory()

if not invObj.execFail:
    invObj.run()

